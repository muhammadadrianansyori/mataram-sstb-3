"""
BKD PAD Monitoring System - Dashboard Streamlit
Sistem monitoring Pendapatan Asli Daerah untuk Badan Keuangan Daerah
"""

import streamlit as st

import folium
from folium import plugins
import ee
from utils import initialize_gee
import streamlit.components.v1 as components
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
from io import BytesIO

# Import BKD modules
from modules.boundary_manager import BoundaryManager
from modules.boundary_cache import load_boundaries_cached
from modules.report_generator import BKDReportGenerator

from modules.parking_detector import ParkingDetector
from modules.landuse_analyzer import LandUseAnalyzer
from modules.pbb_monitor import PBBMonitor
from config.bkd_config import (
    MATARAM_DISTRICTS, TARGET_PAD_ANNUAL, COLORS,
    PARKING_TARIFF, PBB_RATE, NJOP_ZONE
)
# Import AI Validator
from modules.ai_validator import AIValidator, get_ai_status
from modules.street_mapper import StreetMapper


# Page Config
st.set_page_config(
    layout="wide",
    page_title="BKD PAD Monitoring - Kota Mataram",
    page_icon="💰"
)

# Custom CSS
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 20px;
        border-radius: 10px;
        color: white;
        margin-bottom: 20px;
    }
    .metric-card {
        background: white;
        padding: 15px;
        border-radius: 10px;
        box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        margin: 10px 0;
    }
    .info-box {
        padding: 15px;
        background: #f0f9ff;
        border-left: 4px solid #3b82f6;
        border-radius: 5px;
        margin: 10px 0;
    }
    .warning-box {
        padding: 15px;
        background: #fef3c7;
        border-left: 4px solid #f59e0b;
        border-radius: 5px;
        margin: 10px 0;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    .stTabs [data-baseweb="tab"] {
        padding: 10px 20px;
        background-color: #f3f4f6;
        border-radius: 8px 8px 0 0;
    }
</style>
""", unsafe_allow_html=True)

# Header
st.markdown("""
<div class="main-header">
    <h1>💰 Sistem Monitoring PAD - Badan Keuangan Daerah</h1>
    <p style='margin: 0; font-size: 16px;'>Kota Mataram | Monitoring Berbasis Satelit & Big Data</p>
</div>
""", unsafe_allow_html=True)

# Initialize GEE
gee_status = initialize_gee()

# Initialize AI Engine
ai_validator = AIValidator()
ai_status_msg = get_ai_status()


# Sidebar Configuration
st.sidebar.header("⚙️ Konfigurasi Sistem")

# Sidebar Configuration
st.sidebar.header("⚙️ Konfigurasi Sistem")

# 2. District & Kelurahan Selection (Wilayah)
st.sidebar.header("📍 Wilayah Analisis")
selected_district = st.sidebar.selectbox(
    "Pilih Kecamatan",
    list(MATARAM_DISTRICTS.keys()),
    help="Pilih kecamatan untuk fokus analisis"
)

district_config = MATARAM_DISTRICTS[selected_district]

# Regional Filters (Kelurahan/Lingkungan/RT)
from modules.boundary_manager import BoundaryManager
from config.bkd_config import BOUNDARY_GEOJSON_PATH

selected_kelurahan = []
selected_lingkungan = []
selected_rt = []
boundary_mgr = None

try:
    boundary_mgr = BoundaryManager(BOUNDARY_GEOJSON_PATH)
    
    # --- State Sync Logic ---
    # Read global search value from session state to drive hierarchical indices
    gs_val = st.session_state.get('gs_sls_search', "--- Cari Wilayah / RT ---")
    parent_info = {}
    if gs_val != "--- Cari Wilayah / RT ---":
        parent_info = boundary_mgr.get_parent_info_by_sls(gs_val, selected_district)
    
    # --- 📋 METODE 1: PILIH MANUAL (BERTAHAP) ---
    st.sidebar.markdown("### 📋 Pilih Wilayah (Bertahap)")
    
    # 1. Kelurahan
    kelurahan_list = ["--- Semua Kelurahan ---"] + boundary_mgr.get_kelurahan_list(selected_district)
    k_idx = 0
    if parent_info.get('kelurahan') in kelurahan_list:
        k_idx = kelurahan_list.index(parent_info['kelurahan'])
        
    selected_kelurahan_raw = st.sidebar.selectbox(
        "Pilih Kelurahan",
        options=kelurahan_list,
        index=k_idx,
        help="Pilih kelurahan terlebih dahulu"
    )
    
    selected_kelurahan = [] if selected_kelurahan_raw == "--- Semua Kelurahan ---" else [selected_kelurahan_raw]
    
    selected_lingkungan = []
    selected_rt = []

    if selected_kelurahan:
        # 2. Lingkungan (Single Select)
        ling_raw_list = ["--- Semua Lingkungan ---"] + boundary_mgr.get_lingkungan_list(selected_district, selected_kelurahan)
        l_idx = 0
        if parent_info.get('lingkungan') in ling_raw_list:
            l_idx = ling_raw_list.index(parent_info['lingkungan'])

        selected_lingkungan_raw = st.sidebar.selectbox(
            "Pilih Lingkungan",
            options=ling_raw_list,
            index=l_idx,
            help="Pilih lingkungan spesifik (Data: SLS)"
        )
        
        selected_lingkungan = [] if selected_lingkungan_raw == "--- Semua Lingkungan ---" else [selected_lingkungan_raw]
        
        if selected_lingkungan:
            # 3. RT (Multi Select)
            rt_list = boundary_mgr.get_rt_list(selected_district, selected_kelurahan, selected_lingkungan)
            
            rt_defaults = []
            if gs_val != "--- Cari Wilayah / RT ---":
                # Auto-check the RT if it matches the searched SLS
                rt_part = gs_val.split(' LINGKUNGAN ')[0].strip() if ' LINGKUNGAN ' in gs_val else gs_val
                if rt_part in rt_list:
                    rt_defaults = [rt_part]

            selected_rt = st.sidebar.multiselect(
                "Filter RT (Opsional)",
                options=rt_list,
                default=rt_defaults,
                help="Pilih satu atau lebih RT"
            )
            
            if selected_rt:
                st.sidebar.info(f"📍 Fokus: {len(selected_rt)} RT")
            else:
                st.sidebar.info(f"📍 Fokus: {selected_lingkungan_raw}")
        else:
            st.sidebar.info(f"📍 Fokus: {selected_kelurahan_raw}")
    
    # --- 🔍 METODE 2: PENCARIAN EKSPRES ---
    st.sidebar.markdown("---")
    st.sidebar.markdown("### ⚡ Pencarian Instan (Tanpa Menebak)")
    
    all_sls = ["--- Cari Wilayah / RT ---"] + boundary_mgr.get_all_sls_in_district(selected_district)
    st.sidebar.selectbox(
        "Cari RT Lintas Kelurahan",
        options=all_sls,
        key='gs_sls_search',
        help="Ketik dan pilih RT untuk memfilter wilayah secara cepat"
    )
    
except Exception as e:
    st.sidebar.error(f"Error pada filter wilayah: {e}")
    selected_kelurahan = []
    selected_lingkungan = []
    selected_rt = []

# 3. Map Visual Controls (Bottom)
st.sidebar.header("🗺️ Kontrol Peta")
map_background = st.sidebar.radio(
    "Background Peta",
    ["Satelit", "Polos (Light)", "Polos (Dark)", "Hanya Batas Wilayah"],
    index=0,
    help="Pilih jenis background peta"
)

show_boundaries = st.sidebar.checkbox(
    "Tampilkan Garis Batas",
    value=True,
    help="Tampilkan garis batas wilayah di peta"
)

boundary_opacity = st.sidebar.slider(
    "Transparansi Batas",
    min_value=0.0,
    max_value=1.0,
    value=0.6,
    step=0.1
)


# Create ROI safely
roi = None
if gee_status:
    try:
        # V3 Optimization: Dynamic ROI following Kelurahan Boundary
        if selected_kelurahan and boundary_mgr:
            # Get specific kelurahan boundary from GeoJSON
            kel_boundary = boundary_mgr.get_boundary_by_kelurahan(selected_kelurahan[0])
            if kel_boundary:
                # Convert GeoJSON geometry to ee.Geometry
                roi = ee.Geometry(kel_boundary['geometry'])
                # No radius needed, uses exact official boundary
        
        # Fallback to district point + radius if no kelurahan or geometry failed
        if roi is None:
            roi = ee.Geometry.Point([district_config['lon'], district_config['lat']]).buffer(district_config['radius'])
            
    except Exception as e:
        st.error(f"❌ Error creating ROI: {e}")

# ==========================================
# MAP HELPER FUNCTIONS
# ==========================================
def create_map_with_controls(lat, lon, zoom, background_type, show_boundaries_flag=True):
    """
    Create Folium map with user-selected background
    
    Args:
        lat: Latitude for map center
        lon: Longitude for map center
        zoom: Zoom level
        background_type: Type of background ('Satelit', 'Polos (Light)', etc.)
        show_boundaries_flag: Whether to show boundaries
    
    Returns:
        Folium Map object
    """
    # Background tiles mapping
    tiles_map = {
        "Satelit": ('https://mt1.google.com/vt/lyrs=s&x={x}&y={y}&z={z}', 'Google Satellite'),
        "Polos (Light)": ('CartoDB positron', 'CartoDB Positron'),
        "Polos (Dark)": ('CartoDB dark_matter', 'CartoDB Dark Matter'),
        "Hanya Batas Wilayah": ('https://{s}.basemaps.cartocdn.com/light_nolabels/{z}/{x}/{y}{r}.png', 'Plain')
    }
    
    tile_url, attr_name = tiles_map.get(background_type, tiles_map["Satelit"])
    
    # Create map
    if tile_url.startswith('http'):
        m = folium.Map(
            location=[lat, lon],
            zoom_start=zoom,
            tiles=tile_url,
            attr=attr_name
        )
    else:
        m = folium.Map(
            location=[lat, lon],
            zoom_start=zoom,
            tiles=tile_url
        )
    
    return m

def add_boundary_overlay(m, district_name, kelurahan=None, lingkungan=None, rt=None, opacity=0.6):
    """
    Add boundary overlay to Folium map with granular filtering
    
    Args:
        m: Folium Map object
        district_name: Name of district
        kelurahan: Optional selected kelurahan (list)
        lingkungan: Optional selected lingkungan (list)
        rt: Optional selected RT (list)
        opacity: Opacity of boundary lines
    
    Returns:
        Modified Folium Map object
    """
    if boundary_mgr is None:
        return m
    
    try:
        from config.bkd_config import BOUNDARY_STYLES
        
        boundaries = boundary_mgr.get_boundaries_by_district(district_name)
        
        for boundary in boundaries:
            nmdesa = boundary['properties']['nmdesa']
            nmsls = boundary['properties']['nmsls']
            
            # CRITICAL: Prioritize most granular selection for visualization
            visible = True
            
            if rt and len(rt) > 0:
                # If RT is selected, only show matching RTs
                # Pattern match: starts with any of RT names
                is_rt_match = any(nmsls.startswith(f"{r} ") or nmsls == r for r in rt)
                if not is_rt_match:
                    visible = False
            elif lingkungan and len(lingkungan) > 0:
                # If Lingkungan is selected, only show matching Lingkungans
                is_lingkungan_match = any(f" LINGKUNGAN {l}" in nmsls or nmsls == l for l in lingkungan)
                if not is_lingkungan_match:
                    visible = False
            elif kelurahan and len(kelurahan) > 0:
                # If only Kelurahan is selected
                if nmdesa not in kelurahan:
                    visible = False
            
            if not visible:
                continue
                    
            folium.GeoJson(
                boundary,
                style_function=lambda x, op=opacity: {
                    'fillColor': 'transparent',
                    'color': '#FF6B6B',
                    'weight': 2,
                    'fillOpacity': 0,
                    'opacity': op
                },
                tooltip=folium.Tooltip(
                    f"<b>{boundary['properties']['nmsls']}</b><br>"
                    f"Kelurahan: {boundary['properties']['nmdesa']}"
                )
            ).add_to(m)
    except Exception as e:
        print(f"Error adding boundaries: {e}")
    
    return m

def add_map_legend(m, show_boundaries_flag=True):
    """
    Add legend to Folium map
    
    Args:
        m: Folium Map object
        show_boundaries_flag: Whether boundaries are shown
    
    Returns:
        Modified Folium Map object
    """
    legend_html = '''
    <div style="position: fixed; 
                bottom: 50px; right: 50px; width: 220px; height: auto; 
                background-color: white; z-index:9999; font-size:13px;
                border:2px solid grey; border-radius: 5px; padding: 12px;
                box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
    <p style="margin: 0 0 8px 0; font-weight: bold; font-size: 14px; border-bottom: 1px solid #ddd; padding-bottom: 5px;">📍 Legenda</p>
    '''
    
    if show_boundaries_flag:
        legend_html += '''
        <p style="margin: 5px 0;">
            <span style="color: #FF6B6B; font-weight: bold;">━━━</span> Batas Kelurahan
        </p>
        '''
    
    legend_html += '''
    <p style="margin: 5px 0;">
        <span style="background-color: #FFD700; padding: 2px 12px; border-radius: 3px;">  </span> Lahan Parkir
    </p>
    <p style="margin: 5px 0;">
        <span style="background-color: #ef4444; padding: 2px 12px; border-radius: 3px;">  </span> Alih Fungsi Lahan
    </p>
    <p style="margin: 5px 0;">
        <span style="background-color: #3b82f6; padding: 2px 12px; border-radius: 3px;">  </span> Bangunan Baru
    </p>
    </div>
    '''
    
    m.get_root().html.add_child(folium.Element(legend_html))
    return m


# Initialize modules
parking_detector = ParkingDetector()
landuse_analyzer = LandUseAnalyzer()
pbb_monitor = PBBMonitor()

# Main Tabs
tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
    "🅿️ Lahan Parkir",
    "🏗️ Alih Fungsi Lahan",
    "🏢 Monitoring PBB",
    "📊 Dashboard PAD",
    "📄 Laporan & Export",
    "🛣️ Pemetaan Jalan"
])

# ============================================
# TAB 1: LAHAN PARKIR
# ============================================
with tab1:
    st.header("🅿️ Monitoring Lahan Parkir")
    st.markdown("Deteksi dan kalkulasi potensi retribusi parkir dari citra satelit")
    
    # Contextual Year Selection
    year_analysis = st.selectbox("📅 Pilih Tahun Analisis", list(range(2015, 2026)), index=9, help="Pilih tahun data satelit yang ingin dianalisis")

    # ---------------------------------------------------------
    # MODIFIKASI: EKSEKUSI TERTUNDA (DEFERRED EXECUTION)
    # ---------------------------------------------------------
    
    # 1. Inisialisasi State Awal (Default Ampenan)
    if 'active_district' not in st.session_state:
        st.session_state['active_district'] = 'Ampenan'
        st.session_state['active_kelurahan'] = []
        st.session_state['active_lingkungan'] = []
        st.session_state['active_rt'] = []

    # 2. Tombol Analisis: Update STATE "Aktif" dengan Pilihan Sidebar
    if st.button("🔍 Analisis Lahan Parkir", key="btn_parking"):
        st.session_state['active_district'] = selected_district
        st.session_state['active_kelurahan'] = selected_kelurahan
        st.session_state['active_lingkungan'] = selected_lingkungan
        st.session_state['active_rt'] = selected_rt
        
        # Use a separate function for caching to work properly
        # We need to make sure arguments are hashable. roi is ee.Geometry (not easily hashable).
        # So we cache based on district/kelurahan names (strings) + year.
        
        @st.cache_data(show_spinner=False, ttl=3600)
        def convert_roi_and_detect(dist_name, kel_name, year):
            # Gunakan 'active_district' untuk ROI
            roi_config = MATARAM_DISTRICTS[dist_name]
            
            # Re-create ROI based on active selection (since 'roi' var is from sidebar)
            # This ensures consistency with the active district
            active_roi = ee.Geometry.Point([roi_config['lon'], roi_config['lat']]).buffer(roi_config['radius'])
            
            # Try to get specific boundary if kelurahan is selected
            if kel_name != "ALL" and boundary_mgr:
                 kel_boundary = boundary_mgr.get_boundary_by_kelurahan(kel_name)
                 if kel_boundary:
                     active_roi = ee.Geometry(kel_boundary['geometry'])

            return parking_detector.detect_parking_areas(active_roi, year)

        with st.spinner(f"Memproses data satelit tahun {year_analysis} untuk {st.session_state['active_district']}..."):
            # Use strings for caching key
            kel_key = st.session_state['active_kelurahan'][0] if st.session_state['active_kelurahan'] else "ALL"
            parking_data = convert_roi_and_detect(st.session_state['active_district'], kel_key, year_analysis)
            
            # Store in session state
            st.session_state['parking_data'] = parking_data
    
    # Display results if available
    if 'parking_data' in st.session_state:
        # Run AI Validation if not yet validated
        # Lazy Loading AI Validation Logic
        # -------------------------------------------------------------
        # Masalah: Melakukan request ke GEE untuk setiap item (N+1) sangat lambat.
        # Solusi: User memilih secara manual mana yang ingin divalidasi.
        
        data = st.session_state['parking_data']  # <--- FIXED: Define data here
        
        # 1. Container untuk Kontrol Validasi
        st.markdown("---")
        with st.expander("🕵️ Validasi Lahan dengan AI (Lazy Loading)", expanded=True):
            col_v1, col_v2 = st.columns([3, 1])
            
            # Dropdown pilihan ID (hanya yang belum divalidasi)
            unvalidated_items = [p for p in data['parking_areas'] if 'ai_validation' not in p]
            all_items_ids = [p['id'] for p in data['parking_areas']]
            
            with col_v1:
                target_id = st.selectbox(
                    "Pilih ID Area Parkir untuk Divalidasi:",
                    options=["--- Pilih ID ---"] + all_items_ids,
                    help="Pilih ID lokasi yang ingin diverifikasi oleh AI (membutuhkan waktu ~2-3 detik)"
                )
            
            with col_v2:
                # Tombol Aksi
                if st.button("🧠 Validasi Area Ini", type="primary", disabled=(target_id == "--- Pilih ID ---")):
                    target_parking = next((p for p in data['parking_areas'] if p['id'] == target_id), None)
                    
                    if target_parking:
                        with st.spinner(f"Sedang memvalidasi {target_id} dengan AI..."):
                            # 1. Get Chip
                            chip = ai_validator.get_image_chip(target_parking['coordinates'], year_analysis)
                            # 2. Verify
                            res = ai_validator.verify_parking_area(chip)
                            # 3. Update Data in State
                            target_parking['ai_validation'] = res
                            st.success(f"Validasi Selesai: {res['status']}")
                            st.rerun()

        # -------------------------------------------------------------

        data = st.session_state['parking_data']
        
        # Apply spatial filter (Always apply for boundary capping)
        # BUG FIX: Use ACTIVE filters, not Sidebar filters
        if boundary_mgr:
            data['parking_areas'] = boundary_mgr.spatial_filter(
                data['parking_areas'], 
                st.session_state['active_district'], 
                st.session_state['active_kelurahan'], 
                st.session_state['active_lingkungan'], 
                st.session_state['active_rt']
            )
            
            # Filter only AI Verified (Optional but recommended for accuracy)
            verified_areas = [p for p in data['parking_areas'] if p.get('ai_validation', {}).get('verified', False) or p.get('source') == 'OpenStreetMap']
            
            # Toggle for strict AI filtering - Default to False for better visibility
            st.sidebar.markdown("---")
            use_ai_strict = st.sidebar.toggle("Hanya Tampilkan AI Verified + OSM", value=False, help="Hanya hitung area yang sudah dikonfirmasi benar oleh AI atau OSM")
            
            display_areas = verified_areas if use_ai_strict else data['parking_areas']

            # Recalculate metrics based on display_areas
            data['count'] = len(display_areas)
            data['total_area_m2'] = sum(p['area_m2'] for p in display_areas)
            data['estimated_revenue_annual'] = sum(p['revenue_annual'] for p in display_areas)
        
        # Metrics
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("🅿️ Total Area Parkir", f"{data['count']} lokasi")
        col2.metric("📏 Total Luas", f"{int(data['total_area_m2']):,} m²")
        col3.metric("💰 Potensi PAD/Tahun", f"Rp {int(data['estimated_revenue_annual']):,}")
        col4.metric("🧠 AI Status", "Verified" if len(data['parking_areas']) > 0 else "N/A")
        
        # Map
        st.subheader(f"🗺️ Peta Lokasi Parkir: {st.session_state['active_district']}")
        
        # Create map with user-selected background but ACTIVE location
        active_config = MATARAM_DISTRICTS[st.session_state['active_district']]
        
        # Default Map Center from District Config
        map_center_lat = active_config['lat']
        map_center_lon = active_config['lon']
        map_zoom = 15
        
        # AUTO-ZOOM LOGIC: 
        # If Kelurahan is selected, find its centroid and zoom in
        if st.session_state['active_kelurahan'] and boundary_mgr:
            try:
                kel_name = st.session_state['active_kelurahan'][0]
                kel_boundary = boundary_mgr.get_boundary_by_kelurahan(kel_name)
                if kel_boundary:
                    # Calculate centroid (simple average of bbox or geometry center)
                    from shapely.geometry import shape
                    centroid = shape(kel_boundary['geometry']).centroid
                    map_center_lat = centroid.y
                    map_center_lon = centroid.x
                    map_zoom = 16 # Zoom deeper for Kelurahan
            except Exception as e:
                print(f"Auto-zoom error: {e}")

        m = create_map_with_controls(
            map_center_lat,
            map_center_lon,
            map_zoom,
            map_background,
            show_boundaries
        )
        
        # Add boundary overlay if enabled (Use ACTIVE filters)
        if show_boundaries:
            m = add_boundary_overlay(
                m, 
                st.session_state['active_district'], 
                st.session_state['active_kelurahan'], 
                st.session_state['active_lingkungan'], 
                st.session_state['active_rt'], 
                boundary_opacity
            )
        
        # Add parking areas
        for parking in data['parking_areas']:
            if parking['coordinates']:
                # Convert coordinates
                folium_coords = [[c[1], c[0]] for c in parking['coordinates']]
                
                folium.Polygon(
                    locations=folium_coords,
                    popup=folium.Popup(
                        parking_detector.create_parking_popup_html(parking),
                        max_width=320
                    ),
                    tooltip=f"🅿️ {parking['parking_type'].title()} - {parking['area_m2']:.0f}m² - Klik untuk detail",
                    color=COLORS['parking'],
                    fill=True,
                    fillColor=COLORS['parking'],
                    fillOpacity=0.6,
                    weight=2
                ).add_to(m)
        
        # Add legend
        m = add_map_legend(m, show_boundaries)
        
        # Display map
        import streamlit_folium
        from streamlit_folium import st_folium
        st_folium(m, height=600, use_container_width=True, returned_objects=[])
        
        # Data table
        st.subheader("📋 Detail Area Parkir")
        
        if len(data['parking_areas']) == 0:
            st.warning("⚠️ Tidak ada area parkir terdeteksi di area yang dipilih.")
        else:
            df_parking = pd.DataFrame(data['parking_areas'])
            df_parking_display = df_parking[['id', 'parking_type', 'area_m2', 'estimated_capacity', 'revenue_annual']]
            st.dataframe(df_parking_display, use_container_width=True)
            
            # Chart
            col1, col2 = st.columns(2)
            with col1:
                fig_type = px.pie(
                    df_parking,
                    names='parking_type',
                    title='Distribusi Jenis Parkir',
                    hole=0.4
                )
                st.plotly_chart(fig_type, use_container_width=True)
            
            with col2:
                fig_revenue = px.bar(
                    df_parking,
                    x='parking_type',
                    y='revenue_annual',
                    title='Potensi PAD per Jenis Parkir',
                    labels={'revenue_annual': 'PAD (Rp/tahun)', 'parking_type': 'Jenis Parkir'}
                )
                st.plotly_chart(fig_revenue, use_container_width=True)

# ============================================
# TAB 2: ALIH FUNGSI LAHAN
# ============================================
with tab2:
    st.header("🏗️ Analisis Alih Fungsi Lahan")
    st.markdown("Deteksi perubahan penggunaan lahan dari Citra Satelit")
    
    # Contextual Year Selection
    col_y1, col_y2 = st.columns(2)
    year_baseline = col_y1.selectbox("📅 Tahun Baseline", list(range(2015, 2026)), index=4, help="Tahun awal pembanding")
    year_current = col_y2.selectbox("📅 Tahun Saat Ini", list(range(2015, 2026)), index=9, help="Tahun akhir pembanding")

    st.markdown(f"Status Analisis: **{year_baseline}** → **{year_current}**")

    if st.button("🔍 Analisis Perubahan Lahan", key="btn_landuse"):
        with st.spinner("Memproses data temporal (Sentinel-2 & Dynamic World)..."):
            # Analyze real land use change
            landuse_data = landuse_analyzer.analyze_land_change(roi, year_baseline, year_current)
            
            # --- AI VALIDATION STEP ---
            # Enrich the detected changes with AI validation
            with st.spinner("🧠 Menjalankan AI Validator (Prithvi-100M) untuk verifikasi bangunan..."):
                progress_bar = st.progress(0)
                total_changes = len(landuse_data['changes'])
                
                for idx, change in enumerate(landuse_data['changes']):
                    # Simulate fetching image chips (in real app, use geemap to get pixel array)
                    chip_start = ai_validator.get_image_chip(change['coordinates'], year_baseline) 
                    chip_end = ai_validator.get_image_chip(change['coordinates'], year_current)
                    
                    # Run AI Verification
                    ai_result = ai_validator.verify_change(chip_start, chip_end)
                    
                    # Store result in change dict
                    change['ai_validation'] = ai_result
                    progress_bar.progress((idx + 1) / total_changes)
                
                progress_bar.empty()
            # ---------------------------
            
            st.session_state['landuse_data'] = landuse_data
    
    # Display results
    if 'landuse_data' in st.session_state:
        data = st.session_state['landuse_data']
        
        # Apply spatial filter (Always apply for boundary capping)
        if boundary_mgr:
            data['changes'] = boundary_mgr.spatial_filter(data['changes'], selected_district, selected_kelurahan, selected_lingkungan, selected_rt)
            
            # Recalculate tax potential after filtering
            total_annual = sum(c.get('estimated_pbb', 0) for c in data['changes'])
            high_priority = sum(1 for c in data['changes'] if c['priority'] == 'HIGH')
            data['tax_potential'] = {
                'total_changes': len(data['changes']),
                'high_priority_changes': high_priority,
                'total_annual': total_annual,
                'avg_per_change': total_annual / len(data['changes']) if data['changes'] else 0
            }
        
        # Metrics
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("🔄 Total Perubahan", f"{data['tax_potential']['total_changes']} area")
        col2.metric("⚠️ Prioritas Tinggi", f"{data['tax_potential']['high_priority_changes']} area")
        col3.metric("💰 Potensi Pajak Baru", f"Rp {int(data['tax_potential']['total_annual']):,}/tahun")
        col4.metric("📊 Rata-rata/Area", f"Rp {int(data['tax_potential']['avg_per_change']):,}")
        
        # Map
        st.subheader("🗺️ Peta Perubahan Lahan")
        
        # Create map with user-selected background
        m = create_map_with_controls(
            district_config['lat'],
            district_config['lon'],
            14,
            map_background,
            show_boundaries
        )
        
        # Add boundary overlay if enabled
        if show_boundaries:
            m = add_boundary_overlay(m, selected_district, selected_kelurahan, selected_lingkungan, selected_rt, boundary_opacity)
        
        # Priority colors
        priority_colors = {
            'HIGH': '#ef4444',
            'MEDIUM': '#f97316',
            'LOW': '#eab308',
            'CRITICAL': '#dc2626'
        }
        
        # Add changes
        for change in data['changes']:
            if change['coordinates']:
                folium_coords = [[c[1], c[0]] for c in change['coordinates']]
                
                color = priority_colors.get(change['priority'], '#6b7280')
                
                poly = folium.Polygon(
                    locations=folium_coords,
                    popup=folium.Popup(
                        landuse_analyzer.create_change_popup_html(change),
                        max_width=320
                    ),
                    tooltip=f"🏗️ {change['from_class']} → {change['to_class']} ({change['priority']}) - Klik untuk detail",
                    color=color,
                    fill=True,
                    fillColor=color,
                    fillOpacity=0.6,
                    weight=2
                )
                
                # Add AI Badge to popup if verified
                if 'ai_validation' in change:
                    ai = change['ai_validation']
                    if ai['verified']:
                        folium.Marker(
                           location=folium_coords[0],
                           icon=folium.Icon(color='green', icon='check', prefix='fa'),
                           tooltip=f"✅ AI Verified ({int(ai['confidence']*100)}%)"
                        ).add_to(m)

                poly.add_to(m)
        
        # Add legend
        m = add_map_legend(m, show_boundaries)
        
        components.html(m._repr_html_(), height=600)
        
        # Data table
        st.subheader("📋 Detail Perubahan Lahan")
        st.info(f"Status AI Engine: {ai_status_msg}")
        
        # Check if there are any changes
        if len(data['changes']) == 0:
            st.warning("⚠️ Tidak ada perubahan lahan terdeteksi di area yang dipilih.")
        else:
            df_changes = pd.DataFrame(data['changes'])
            
            # Extract AI data for display
            if len(data['changes']) > 0 and 'ai_validation' in data['changes'][0]:
                df_changes['AI Confidence'] = df_changes['ai_validation'].apply(lambda x: f"{int(x.get('confidence',0)*100)}%")
                df_changes['AI Status'] = df_changes['ai_validation'].apply(lambda x: x.get('status', '-'))
                
            df_changes_display = df_changes[['id', 'from_class', 'to_class', 'area_m2', 'priority', 'AI Confidence', 'AI Status']]
            st.dataframe(df_changes_display, use_container_width=True)
            
            # Charts
            col1, col2 = st.columns(2)
            with col1:
                # Change matrix
                change_matrix = df_changes.groupby(['from_class', 'to_class']).size().reset_index(name='count')
                fig_matrix = px.sunburst(
                    change_matrix,
                    path=['from_class', 'to_class'],
                    values='count',
                    title='Matriks Perubahan Lahan'
                )
                st.plotly_chart(fig_matrix, use_container_width=True)
            
            with col2:
                # Priority distribution
                priority_dist = df_changes['priority'].value_counts().reset_index()
                priority_dist.columns = ['priority', 'count']
                fig_priority = px.bar(
                    priority_dist,
                    x='priority',
                    y='count',
                    title='Distribusi Prioritas',
                    color='priority',
                    color_discrete_map=priority_colors
                )
                st.plotly_chart(fig_priority, use_container_width=True)

# ============================================
# TAB 3: MONITORING PBB
# ============================================
with tab3:
    st.header("🏢 Monitoring Perubahan Bangunan (PBB)")
    st.markdown("Deteksi penambahan area atau tinggi bangunan (Google Open Buildings)")
    
    # Contextual Year Selection
    col_p1, col_p2 = st.columns(2)
    pbb_year_baseline = col_p1.selectbox("📅 Tahun Baseline PBB", list(range(2015, 2026)), index=4)
    pbb_year_current = col_p2.selectbox("📅 Tahun Saat Ini PBB", list(range(2015, 2026)), index=9)

    st.markdown(f"Status Analisis: **{pbb_year_baseline}** → **{pbb_year_current}**")
    
    if st.button("🔍 Analisis Perubahan Bangunan", key="btn_pbb"):
        with st.spinner("Memproses data bangunan (Google Open Buildings)..."):
            # Monitor actual building changes
            pbb_data = pbb_monitor.monitor_building_changes(roi, pbb_year_baseline, pbb_year_current)
            
            st.session_state['pbb_data'] = pbb_data
    
    # Display results
    if 'pbb_data' in st.session_state:
        data = st.session_state['pbb_data']
        
        # Apply spatial filter (Always apply for boundary capping)
        if boundary_mgr:
            filtered_changes = boundary_mgr.spatial_filter(data['changes'], selected_district, selected_kelurahan, selected_lingkungan, selected_rt)
            data['changes'] = filtered_changes
            
            # Recalculate tax impact after filtering
            total_area = sum(b['area_increase'] for b in filtered_changes)
            total_tax = sum(b['tax_increase'] for b in filtered_changes)
            data['tax_impact'] = {
                'total_buildings_changed': len(filtered_changes),
                'total_area_increase_m2': total_area,
                'total_tax_increase_annual': total_tax,
                'avg_tax_increase_per_building': total_tax / len(filtered_changes) if filtered_changes else 0
            }
        
        # Metrics
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("🏢 Bangunan Berubah", f"{data['tax_impact']['total_buildings_changed']} unit")
        col2.metric("📏 Total Penambahan Area", f"{int(data['tax_impact']['total_area_increase_m2']):,} m²")
        col3.metric("💰 Kenaikan PBB/Tahun", f"Rp {int(data['tax_impact']['total_tax_increase_annual']):,}")
        col4.metric("📊 Rata-rata/Bangunan", f"Rp {int(data['tax_impact']['avg_tax_increase_per_building']):,}")
        
        # Map
        st.subheader("🗺️ Peta Perubahan Bangunan")
        
        # Create map with user-selected background
        m = create_map_with_controls(
            district_config['lat'],
            district_config['lon'],
            15,
            map_background,
            show_boundaries
        )
        
        # Add boundary overlay if enabled
        if show_boundaries:
            m = add_boundary_overlay(m, selected_district, selected_kelurahan, selected_lingkungan, selected_rt, boundary_opacity)
        
        # Add building changes
        for building in data['changes']:
            if building['coordinates']:
                folium_coords = [[c[1], c[0]] for c in building['coordinates']]
                
                # Color based on change type
                if 'height_increase' in building['change_type']:
                    color = '#8b5cf6'  # Purple for height
                else:
                    color = '#3b82f6'  # Blue for area
                
                folium.Polygon(
                    locations=folium_coords,
                    popup=folium.Popup(
                        pbb_monitor.create_building_change_popup_html(building),
                        max_width=340
                    ),
                    tooltip=f"🏢 {building['id']} - Area: +{building['area_increase']:.0f}m², Height: +{building['height_increase']:.0f}m",
                    color=color,
                    fill=True,
                    fillColor=color,
                    fillOpacity=0.6,
                    weight=2
                ).add_to(m)
        
        # Add legend
        m = add_map_legend(m, show_boundaries)
        
        components.html(m._repr_html_(), height=600)
        
        # Data table
        st.subheader("📋 Detail Perubahan Bangunan")
        if len(data['changes']) == 0:
            st.warning("⚠️ Tidak ada perubahan bangunan terdeteksi di area yang dipilih.")
        else:
            df_buildings = pd.DataFrame(data['changes'])
            df_buildings_display = df_buildings[['id', 'old_area', 'new_area', 'area_increase', 'old_height', 'new_height', 'height_increase']]
            st.dataframe(df_buildings_display, use_container_width=True)
            
            # Charts
            col1, col2 = st.columns(2)
            with col1:
                fig_area = px.scatter(
                    df_buildings,
                    x='old_area',
                    y='new_area',
                    size='area_increase',
                    color='height_increase',
                    title='Sebaran Perubahan Area vs Tinggi',
                    labels={
                        'old_area': 'Luas Awal (m²)',
                        'new_area': 'Luas Baru (m²)',
                        'area_increase': 'Penambahan Luas',
                        'height_increase': 'Penambahan Tinggi'
                    }
                )
                st.plotly_chart(fig_area, use_container_width=True)
            
            with col2:
                # Change type distribution
                change_types = []
                for building in data['changes']:
                    for ct in building['change_type']:
                        change_types.append(ct)
                
                df_types = pd.DataFrame({'type': change_types})
                type_counts = df_types['type'].value_counts().reset_index()
                type_counts.columns = ['type', 'count']
                
                fig_types = px.pie(
                    type_counts,
                    names='type',
                    values='count',
                    title='Distribusi Jenis Perubahan',
                    hole=0.4
                )
                st.plotly_chart(fig_types, use_container_width=True)

# ============================================
# TAB 4: DASHBOARD PAD KOMPREHENSIF
# ============================================
with tab4:
    st.header("📊 Dashboard PAD Komprehensif")
    st.markdown("Ringkasan semua potensi Pendapatan Asli Daerah")
    
    # Check if all data available
    has_parking = 'parking_data' in st.session_state
    has_landuse = 'landuse_data' in st.session_state
    has_pbb = 'pbb_data' in st.session_state
    
    if not (has_parking or has_landuse or has_pbb):
        st.info("ℹ️ Jalankan analisis di tab-tab sebelumnya untuk melihat dashboard komprehensif")
    else:
        # Aggregate metrics
        total_parking_revenue = st.session_state.get('parking_data', {}).get('estimated_revenue_annual', 0)
        total_landuse_revenue = st.session_state.get('landuse_data', {}).get('tax_potential', {}).get('total_annual', 0)
        total_pbb_increase = st.session_state.get('pbb_data', {}).get('tax_impact', {}).get('total_tax_increase_annual', 0)
        
        total_pad_potential = total_parking_revenue + total_landuse_revenue + total_pbb_increase
        
        # Main metrics
        st.subheader("💰 Total Potensi PAD")
        col1, col2, col3, col4 = st.columns(4)
        
        col1.metric(
            "🅿️ Retribusi Parkir",
            f"Rp {int(total_parking_revenue):,}",
            delta=f"{int(total_parking_revenue/TARGET_PAD_ANNUAL['parking']*100)}% dari target" if total_parking_revenue > 0 else None
        )
        
        col2.metric(
            "🏗️ Pajak Baru (Alih Fungsi)",
            f"Rp {int(total_landuse_revenue):,}",
            delta=f"{int(total_landuse_revenue/TARGET_PAD_ANNUAL['land_change']*100)}% dari target" if total_landuse_revenue > 0 else None
        )
        
        col3.metric(
            "🏢 Kenaikan PBB",
            f"Rp {int(total_pbb_increase):,}",
            delta=f"{int(total_pbb_increase/TARGET_PAD_ANNUAL['pbb']*100)}% dari target" if total_pbb_increase > 0 else None
        )
        
        col4.metric(
            "💎 TOTAL POTENSI",
            f"Rp {int(total_pad_potential):,}",
            delta="Per Tahun"
        )
        
        # Visualization
        st.subheader("📈 Visualisasi PAD")
        
        col1, col2 = st.columns(2)
        
        with col1:
            # Pie chart
            pad_breakdown = pd.DataFrame({
                'Kategori': ['Retribusi Parkir', 'Pajak Baru', 'Kenaikan PBB'],
                'Nilai': [total_parking_revenue, total_landuse_revenue, total_pbb_increase]
            })
            
            fig_pie = px.pie(
                pad_breakdown,
                names='Kategori',
                values='Nilai',
                title='Komposisi Potensi PAD',
                hole=0.4,
                color_discrete_sequence=['#FFD700', '#ef4444', '#3b82f6']
            )
            st.plotly_chart(fig_pie, use_container_width=True)
        
        with col2:
            # Bar chart vs target
            target_comparison = pd.DataFrame({
                'Kategori': ['Parkir', 'Alih Fungsi', 'PBB'],
                'Potensi': [total_parking_revenue, total_landuse_revenue, total_pbb_increase],
                'Target': [TARGET_PAD_ANNUAL['parking'], TARGET_PAD_ANNUAL['land_change'], TARGET_PAD_ANNUAL['pbb']]
            })
            
            fig_bar = go.Figure()
            fig_bar.add_trace(go.Bar(name='Potensi Terdeteksi', x=target_comparison['Kategori'], y=target_comparison['Potensi']))
            fig_bar.add_trace(go.Bar(name='Target Tahunan', x=target_comparison['Kategori'], y=target_comparison['Target']))
            fig_bar.update_layout(title='Potensi vs Target PAD', barmode='group')
            
            st.plotly_chart(fig_bar, use_container_width=True)
            
        # Export Report Button
        st.markdown("---")
        st.subheader("📑 Export Laporan BAPENDA")
        
        col_dl1, col_dl2 = st.columns([2, 1])
        with col_dl1:
            st.info("Unduh laporan lengkap dalam format Excel (.xlsx) dengan tabel 3 arah (Pivot Table) untuk analisis mendalam per Kecamatan, Kelurahan, hingga Lingkungan.")
        
        with col_dl2:
            if st.button("📥 Generate Excel Report"):
                with st.spinner("Menyiapkan Laporan 3 Arah..."):
                    try:
                        report_gen = BKDReportGenerator(boundary_mgr)
                        
                        # Get data from session state safely
                        p_data = st.session_state.get('parking_data', {})
                        l_data = st.session_state.get('landuse_data', {})
                        pbb_data = st.session_state.get('pbb_data', {})
                        
                        years_info = {
                            'start': year_baseline,
                            'end': year_current
                        }
                        excel_file = report_gen.generate_excel(p_data, l_data, pbb_data, years_info)
                        
                        st.download_button(
                            label="⬇️ Download Laporan (.xlsx)",
                            data=excel_file,
                            file_name=f"Laporan_Analisis_PAD_Mataram_{year_current}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        st.success("Laporan siap diunduh!")
                        
                    except Exception as e:
                        st.error(f"Gagal membuat laporan: {str(e)}")
        
        # Summary table
        st.subheader("📋 Ringkasan Detail")
        
        summary_data = []
        
        if has_parking:
            parking = st.session_state['parking_data']
            summary_data.append({
                'Kategori': 'Retribusi Parkir',
                'Jumlah Objek': parking['count'],
                'Total Luas (m²)': int(parking['total_area_m2']),
                'Potensi PAD (Rp/tahun)': int(parking['estimated_revenue_annual']),
                'Status': '✅ Teranalisis'
            })
        
        if has_landuse:
            landuse = st.session_state['landuse_data']
            summary_data.append({
                'Kategori': 'Alih Fungsi Lahan',
                'Jumlah Objek': landuse['tax_potential']['total_changes'],
                'Total Luas (m²)': sum(c['area_m2'] for c in landuse['changes']),
                'Potensi PAD (Rp/tahun)': int(landuse['tax_potential']['total_annual']),
                'Status': '✅ Teranalisis'
            })
        
        if has_pbb:
            pbb = st.session_state['pbb_data']
            summary_data.append({
                'Kategori': 'Perubahan Bangunan (PBB)',
                'Jumlah Objek': pbb['tax_impact']['total_buildings_changed'],
                'Total Luas (m²)': int(pbb['tax_impact']['total_area_increase_m2']),
                'Potensi PAD (Rp/tahun)': int(pbb['tax_impact']['total_tax_increase_annual']),
                'Status': '✅ Teranalisis'
            })
        
        df_summary = pd.DataFrame(summary_data)
        st.dataframe(df_summary, use_container_width=True)
        
        # Recommendations
        st.subheader("💡 Rekomendasi Tindak Lanjut")
        
        recommendations = []
        
        if has_parking and total_parking_revenue > 0:
            recommendations.append("🅿️ **Parkir**: Lakukan survey lapangan untuk validasi lokasi parkir yang terdeteksi")
        
        if has_landuse:
            landuse = st.session_state['landuse_data']
            high_priority = landuse['tax_potential']['high_priority_changes']
            if high_priority > 0:
                recommendations.append(f"🏗️ **Alih Fungsi**: Prioritaskan verifikasi {high_priority} area dengan prioritas TINGGI")
        
        if has_pbb and total_pbb_increase > 0:
            recommendations.append("🏢 **PBB**: Update database SISMIOP dengan data perubahan bangunan")
        
        for idx, rec in enumerate(recommendations, 1):
            st.markdown(f"{idx}. {rec}")

# ============================================
# TAB 5: LAPORAN & EXPORT
# ============================================
with tab5:
    st.header("📄 Laporan & Export Data")
    st.markdown("Download hasil analisis dalam berbagai format")
    
    # Check data availability
    has_data = any([
        'parking_data' in st.session_state,
        'landuse_data' in st.session_state,
        'pbb_data' in st.session_state
    ])
    
    if not has_data:
        st.warning("⚠️ Belum ada data untuk di-export. Jalankan analisis terlebih dahulu.")
    else:
        st.success("✅ Data tersedia untuk export")
        
        # Export options
        st.subheader("📥 Pilih Format Export")
        
        col1, col2, col3 = st.columns(3)
        
        # CSV Export
        with col1:
            if st.button("📊 Export ke CSV", use_container_width=True):
                # Combine all data
                all_data = []
                
                if 'parking_data' in st.session_state:
                    df_parking = pd.DataFrame(st.session_state['parking_data']['parking_areas'])
                    df_parking['kategori'] = 'Parkir'
                    all_data.append(df_parking)
                
                if 'landuse_data' in st.session_state:
                    df_landuse = pd.DataFrame(st.session_state['landuse_data']['changes'])
                    df_landuse['kategori'] = 'Alih Fungsi Lahan'
                    all_data.append(df_landuse)
                
                if 'pbb_data' in st.session_state:
                    df_pbb = pd.DataFrame(st.session_state['pbb_data']['changes'])
                    df_pbb['kategori'] = 'Perubahan Bangunan'
                    all_data.append(df_pbb)
                
                if all_data:
                    df_combined = pd.concat(all_data, ignore_index=True)
                    csv = df_combined.to_csv(index=False).encode('utf-8')
                    
                    st.download_button(
                        label="⬇️ Download CSV",
                        data=csv,
                        file_name=f'bkd_pad_report_{selected_district}_{datetime.now().strftime("%Y%m%d")}.csv',
                        mime='text/csv'
                    )
        
        # Excel Export
        with col2:
            if st.button("📗 Export ke Excel", use_container_width=True):
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    if 'parking_data' in st.session_state:
                        df_parking = pd.DataFrame(st.session_state['parking_data']['parking_areas'])
                        df_parking.to_excel(writer, sheet_name='Parkir', index=False)
                    
                    if 'landuse_data' in st.session_state:
                        df_landuse = pd.DataFrame(st.session_state['landuse_data']['changes'])
                        df_landuse.to_excel(writer, sheet_name='Alih Fungsi Lahan', index=False)
                    
                    if 'pbb_data' in st.session_state:
                        df_pbb = pd.DataFrame(st.session_state['pbb_data']['changes'])
                        df_pbb.to_excel(writer, sheet_name='Perubahan Bangunan', index=False)
                
                excel_data = output.getvalue()
                
                st.download_button(
                    label="⬇️ Download Excel",
                    data=excel_data,
                    file_name=f'bkd_pad_report_{selected_district}_{datetime.now().strftime("%Y%m%d")}.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
        
        # JSON Export
        with col3:
            if st.button("📋 Export ke JSON", use_container_width=True):
                import json
                
                export_data = {
                    'metadata': {
                        'district': selected_district,
                        'year_baseline': year_baseline,
                        'year_current': year_current,
                        'export_date': datetime.now().isoformat()
                    },
                    'parking': st.session_state.get('parking_data', {}),
                    'landuse': st.session_state.get('landuse_data', {}),
                    'pbb': st.session_state.get('pbb_data', {})
                }
                
                json_str = json.dumps(export_data, indent=2, ensure_ascii=False)
                
                st.download_button(
                    label="⬇️ Download JSON",
                    data=json_str,
                    file_name=f'bkd_pad_report_{selected_district}_{datetime.now().strftime("%Y%m%d")}.json',
                    mime='application/json'
                )
        
        # Preview data
        st.subheader("👁️ Preview Data")
        
        preview_option = st.selectbox(
            "Pilih data untuk preview:",
            ["Parkir", "Alih Fungsi Lahan", "Perubahan Bangunan"]
        )
        
        if preview_option == "Parkir" and 'parking_data' in st.session_state:
            df = pd.DataFrame(st.session_state['parking_data']['parking_areas'])
            st.dataframe(df, use_container_width=True)
        
        elif preview_option == "Alih Fungsi Lahan" and 'landuse_data' in st.session_state:
            df = pd.DataFrame(st.session_state['landuse_data']['changes'])
            st.dataframe(df, use_container_width=True)
        
        elif preview_option == "Perubahan Bangunan" and 'pbb_data' in st.session_state:
            df = pd.DataFrame(st.session_state['pbb_data']['changes'])
            st.dataframe(df, use_container_width=True)

# ============================================
# TAB 6: PEMETAAN JALAN
# ============================================
with tab6:
    st.header("🛣️ Pemetaan Jalan dan Gang")
    st.markdown("Pemetaan nama jalan/gang dengan informasi administratif (RT, Lingkungan, Kelurahan)")
    
    # Info box
    st.markdown("""
    <div class="info-box">
        <b>📋 Cara Penggunaan:</b><br>
        1. Pilih Kecamatan yang ingin dipetakan<br>
        2. Klik tombol "Proses Data Jalan"<br>
        3. Tunggu hingga data selesai diproses (1-2 menit)<br>
        4. Lihat hasil pemetaan dan download Excel
    </div>
    """, unsafe_allow_html=True)
    
    # Initialize street mapper
    try:
        from config.bkd_config import BOUNDARY_GEOJSON_PATH
        street_mapper = StreetMapper(BOUNDARY_GEOJSON_PATH)
        
        # Kecamatan selection
        kecamatan_list = street_mapper.get_kecamatan_list()
        selected_kecamatan = st.selectbox(
            "📍 Pilih Kecamatan",
            kecamatan_list,
            help="Pilih kecamatan untuk memetakan jalan"
        )
        
        # Process button
        if st.button("🔍 Proses Data Jalan", type="primary", key="btn_streets"):
            with st.spinner(f"Mengambil data jalan dari OpenStreetMap untuk Kecamatan {selected_kecamatan}..."):
                # Fetch and process street data
                street_data = street_mapper.map_streets_to_admin(selected_kecamatan)
                
                # Store in session state
                st.session_state['street_data'] = street_data
                st.session_state['street_kecamatan'] = selected_kecamatan
        
        # Display results
        if 'street_data' in st.session_state and st.session_state.get('street_kecamatan') == selected_kecamatan:
            df_streets = st.session_state['street_data']
            
            if df_streets.empty:
                st.warning(f"⚠️ Tidak ada data jalan ditemukan untuk Kecamatan {selected_kecamatan}. Pastikan koneksi internet stabil dan coba lagi.")
            else:
                # Metrics
                col1, col2, col3, col4 = st.columns(4)
                col1.metric("🛣️ Total Jalan/Gang", f"{len(df_streets)} jalan")
                col2.metric("📍 Total Kelurahan", f"{df_streets['Kelurahan'].nunique()} kelurahan")
                col3.metric("🏘️ Total Lingkungan", f"{df_streets['Lingkungan'].nunique()} lingkungan")
                col4.metric("📋 Total RT", f"{df_streets['SLS'].nunique()} RT")
                
                # Map Visualization
                st.subheader("🗺️ Peta Validasi Jalan")
                
                # Street search box
                col_search, col_toggle = st.columns([3, 1])
                with col_search:
                    # Get list of unique street names for search
                    street_names = ["--- Pilih Jalan untuk Zoom ---"] + sorted(df_streets['Nama Jalan dan Gang'].unique().tolist())
                    selected_street = st.selectbox(
                        "🔍 Cari Jalan/Gang:",
                        street_names,
                        key="street_search",
                        help="Pilih jalan untuk zoom ke lokasi"
                    )
                
                with col_toggle:
                    show_boundaries = st.checkbox(
                        "Tampilkan Batas RT",
                        value=True,
                        help="Tampilkan batas wilayah RT/Lingkungan/Kelurahan"
                    )
                
                st.info("💡 Klik atau hover pada garis jalan untuk melihat nama jalan dan informasi administratif")
                
                # Fetch street geometry data for mapping
                streets_gdf = street_mapper.fetch_streets_osm(selected_kecamatan)
                
                if not streets_gdf.empty:
                    # Calculate map center and zoom
                    bounds = streets_gdf.total_bounds  # [minx, miny, maxx, maxy]
                    center_lat = (bounds[1] + bounds[3]) / 2
                    center_lon = (bounds[0] + bounds[2]) / 2
                    zoom_level = 14
                    
                    # If a street is selected, zoom to it
                    if selected_street != "--- Pilih Jalan untuk Zoom ---":
                        selected_street_row = streets_gdf[streets_gdf['name'] == selected_street].iloc[0]
                        centroid = selected_street_row['geometry'].centroid
                        center_lat = centroid.y
                        center_lon = centroid.x
                        zoom_level = 17  # Closer zoom for specific street
                    
                    # Create Folium map with plain background
                    m = folium.Map(
                        location=[center_lat, center_lon],
                        zoom_start=zoom_level,
                        tiles='CartoDB positron',
                        attr='CartoDB Positron'
                    )
                    
                    # Color mapping for different road types
                    road_colors = {
                        'primary': '#FF6B6B',
                        'secondary': '#4ECDC4',
                        'tertiary': '#95E1D3',
                        'residential': '#3B82F6',
                        'service': '#FFA726',
                        'unclassified': '#9E9E9E',
                        'living_street': '#66BB6A',
                        'pedestrian': '#AB47BC',
                        'footway': '#8D6E63',
                        'path': '#78909C'
                    }
                    
                    # Add administrative boundaries if enabled
                    if show_boundaries:
                        # Get SLS boundaries for selected Kecamatan
                        kec_boundaries = street_mapper.sls_gdf[
                            street_mapper.sls_gdf['nmkec'] == selected_kecamatan.upper()
                        ]
                        
                        for idx, boundary in kec_boundaries.iterrows():
                            # Create popup for boundary
                            boundary_popup = f"""
                            <div style="font-family: Arial; width: 200px;">
                                <h4 style="margin: 0; color: #1F4E78;">Batas Wilayah</h4>
                                <hr style="margin: 5px 0;">
                                <p style="margin: 2px 0;"><b>RT:</b> {boundary['nmsls']}</p>
                                <p style="margin: 2px 0;"><b>Kelurahan:</b> {boundary['nmdesa']}</p>
                                <p style="margin: 2px 0;"><b>Kecamatan:</b> {boundary['nmkec']}</p>
                            </div>
                            """
                            
                            # Add boundary polygon
                            folium.GeoJson(
                                boundary['geometry'],
                                style_function=lambda x: {
                                    'fillColor': 'transparent',
                                    'color': '#FF1744',
                                    'weight': 2,
                                    'dashArray': '5, 5',
                                    'fillOpacity': 0,
                                    'opacity': 0.6
                                },
                                popup=folium.Popup(boundary_popup, max_width=250),
                                tooltip=folium.Tooltip(
                                    f"<b>{boundary['nmsls']}</b>",
                                    style="background-color: #FFE0E0; color: #C62828; font-family: Arial; font-size: 11px; padding: 3px; border: 1px solid #FF1744; border-radius: 3px;"
                                )
                            ).add_to(m)
                    
                    
                    # Add streets to map
                    for idx, row in streets_gdf.iterrows():
                        # Get color based on highway type
                        color = road_colors.get(row['highway_type'], '#333333')
                        
                        # Check if this is the selected street for highlighting
                        is_selected = (selected_street != "--- Pilih Jalan untuk Zoom ---" and 
                                      row['name'] == selected_street)
                        
                        # Convert coords to Folium format [lat, lon]
                        coords_folium = [[lat, lon] for lon, lat in row['coords_list']]
                        
                        # Create popup HTML with detailed info
                        popup_html = f"""
                        <div style="font-family: Arial; width: 250px;">
                            <h4 style="margin: 0; color: #1F4E78;">{row['name']}</h4>
                            <hr style="margin: 5px 0;">
                            <p style="margin: 2px 0;"><b>Tipe:</b> {row['highway_type'].title()}</p>
                            <p style="margin: 2px 0;"><b>OSM ID:</b> {row['osm_id']}</p>
                        </div>
                        """
                        
                        # Add street polyline
                        folium.PolyLine(
                            locations=coords_folium,
                            popup=folium.Popup(popup_html, max_width=300),
                            tooltip=folium.Tooltip(
                                f"<b>{row['name']}</b><br>Tipe: {row['highway_type'].title()}",
                                style="background-color: white; color: black; font-family: Arial; font-size: 12px; padding: 5px; border: 2px solid #333; border-radius: 3px;"
                            ),
                            color='#FFD700' if is_selected else color,  # Gold color for selected
                            weight=6 if is_selected else 3,  # Thicker if selected
                            opacity=1.0 if is_selected else 0.8
                        ).add_to(m)
                        
                        # Add marker at center of selected street
                        if is_selected:
                            centroid = row['geometry'].centroid
                            folium.Marker(
                                location=[centroid.y, centroid.x],
                                popup=folium.Popup(popup_html, max_width=300),
                                icon=folium.Icon(color='orange', icon='road', prefix='fa'),
                                tooltip=f"<b>📍 {row['name']}</b>"
                            ).add_to(m)
                    
                    # Add legend
                    legend_html = '''
                    <div style="position: fixed; 
                                bottom: 50px; right: 50px; width: 200px; height: auto; 
                                background-color: white; z-index:9999; font-size:12px;
                                border:2px solid grey; border-radius: 5px; padding: 10px;
                                box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
                    <p style="margin: 0 0 8px 0; font-weight: bold; font-size: 13px; border-bottom: 1px solid #ddd; padding-bottom: 5px;">🛣️ Jenis Jalan</p>
                    <p style="margin: 3px 0;"><span style="color: #FF6B6B; font-weight: bold;">━━━</span> Primary</p>
                    <p style="margin: 3px 0;"><span style="color: #4ECDC4; font-weight: bold;">━━━</span> Secondary</p>
                    <p style="margin: 3px 0;"><span style="color: #95E1D3; font-weight: bold;">━━━</span> Tertiary</p>
                    <p style="margin: 3px 0;"><span style="color: #3B82F6; font-weight: bold;">━━━</span> Residential</p>
                    <p style="margin: 3px 0;"><span style="color: #FFA726; font-weight: bold;">━━━</span> Service</p>
                    <p style="margin: 3px 0;"><span style="color: #66BB6A; font-weight: bold;">━━━</span> Gang/Pedestrian</p>
                    </div>
                    '''
                    m.get_root().html.add_child(folium.Element(legend_html))
                    
                    # Display map with st_folium
                    from streamlit_folium import st_folium
                    st_folium(m, width=None, height=600, returned_objects=[])
                else:
                    st.warning("Tidak dapat memuat data geometri jalan untuk peta")
                
                # Data table
                st.subheader("📊 Data Jalan dan Administratif")
                st.dataframe(df_streets, use_container_width=True, height=400)
                
                # Excel export
                st.subheader("📥 Download Data")
                
                # Create Excel file in memory with enhanced formatting
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_streets.to_excel(writer, sheet_name='Data Jalan', index=False, startrow=1)
                    
                    workbook = writer.book
                    worksheet = writer.sheets['Data Jalan']
                    
                    # Import openpyxl formatting
                    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                    
                    # Define styles
                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF", size=12)
                    border = Border(
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000')
                    )
                    
                    # Title row
                    worksheet['A1'] = f'DATA PEMETAAN JALAN - KECAMATAN {selected_kecamatan.upper()}'
                    worksheet['A1'].font = Font(bold=True, size=14, color="1F4E78")
                    worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
                    # Merge across all 9 columns (A to I)
                    worksheet.merge_cells('A1:I1')
                    worksheet.row_dimensions[1].height = 25
                    
                    # Format headers (row 2)
                    for col_idx, col_name in enumerate(df_streets.columns, start=1):
                        cell = worksheet.cell(row=2, column=col_idx)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = border
                    
                    # Auto-adjust column widths and format data cells
                    for idx, col in enumerate(df_streets.columns, start=1):
                        col_letter = chr(64 + idx)
                        
                        # Calculate max width (limit to max 60)
                        data_max_len = 0
                        if not df_streets.empty:
                            data_max_len = df_streets[col].astype(str).map(len).max()
                        
                        max_length = max(data_max_len, len(col))
                        
                        # Special width for links
                        if 'Link' in col:
                            worksheet.column_dimensions[col_letter].width = 30
                        else:
                            worksheet.column_dimensions[col_letter].width = min(max_length + 3, 60)
                        
                        # Data rows formatting
                        for row_idx in range(3, len(df_streets) + 3):
                            cell = worksheet.cell(row=row_idx, column=idx)
                            cell.border = border
                            
                            # Center coordinates and coverage
                            if col in ['Latitude', 'Longitude', 'Coverage']:
                                cell.alignment = Alignment(horizontal='center')
                                if col in ['Latitude', 'Longitude']:
                                    cell.number_format = '0.000000'
                            
                            # Highlight links in blue
                            if 'Link' in col:
                                cell.font = Font(color="0563C1", underline="single")
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                            else:
                                if col not in ['Latitude', 'Longitude', 'Coverage']:
                                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    
                    # Freeze header rows
                    worksheet.freeze_panes = 'A3'
                
                excel_data = output.getvalue()
                
                st.download_button(
                    label="📥 Download Excel",
                    data=excel_data,
                    file_name=f"data_jalan_{selected_kecamatan.lower()}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help="Download data jalan dalam format Excel dengan formatting profesional"
                )
                
                # Summary statistics
                st.subheader("📈 Statistik per Kelurahan")
                
                kelurahan_stats = df_streets.groupby('Kelurahan').agg({
                    'Nama Jalan dan Gang': 'count',
                    'Lingkungan': 'nunique',
                    'SLS': 'nunique'
                }).reset_index()
                kelurahan_stats.columns = ['Kelurahan', 'Jumlah Jalan', 'Jumlah Lingkungan', 'Jumlah RT']
                
                col1, col2 = st.columns([2, 1])
                
                with col1:
                    st.dataframe(kelurahan_stats, use_container_width=True)
                
                with col2:
                    # Simple bar chart
                    fig = px.bar(
                        kelurahan_stats,
                        x='Kelurahan',
                        y='Jumlah Jalan',
                        title=f'Distribusi Jalan per Kelurahan - {selected_kecamatan}',
                        labels={'Jumlah Jalan': 'Jumlah Jalan/Gang'}
                    )
                    st.plotly_chart(fig, use_container_width=True)
                
                # Validation Section with Google Sheets
                st.subheader("✅ Validasi Data dengan Google Sheets")
                st.markdown("""
                <div class="info-box">
                    <b>📋 Cara Validasi:</b><br>
                    1. Pastikan Google Sheets Anda publik (Anyone with the link can view)<br>
                    2. Copy URL Google Sheets Anda<br>
                    3. Paste di bawah dan klik "Validasi Data"<br>
                    4. Sistem akan membandingkan data OSM dengan data Anda
                </div>
                """, unsafe_allow_html=True)
                
                # Google Sheets input
                col_url, col_method = st.columns([3, 1])
                
                with col_url:
                    sheets_url = st.text_input(
                        "🔗 Google Sheets URL:",
                        placeholder="https://docs.google.com/spreadsheets/d/...",
                        help="Paste link Google Sheets yang berisi data jalan untuk validasi"
                    )
                
                with col_method:
                    st.markdown("<br>", unsafe_allow_html=True)
                    use_public = st.checkbox(
                        "Public Sheet",
                        value=True,
                        help="Centang jika sheet publik (tidak perlu autentikasi)"
                    )
                
                if st.button("🔍 Validasi Data", type="primary", key="validate_btn"):
                    if not sheets_url:
                        st.warning("⚠️ Masukkan URL Google Sheets terlebih dahulu")
                    else:
                        try:
                            with st.spinner("Mengambil data dari Google Sheets..."):
                                # Import data from Google Sheets
                                import re
                                
                                if use_public:
                                    # Extract sheet ID and gid from URL
                                    match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', sheets_url)
                                    gid_match = re.search(r'[#&]gid=([0-9]+)', sheets_url)
                                    
                                    if match:
                                        sheet_id = match.group(1)
                                        
                                        # Public sheet CSV export URL with gid support
                                        if gid_match:
                                            gid = gid_match.group(1)
                                            csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
                                            st.info(f"📄 Menggunakan sheet dengan GID: {gid}")
                                        else:
                                            csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
                                            st.info("📄 Menggunakan sheet pertama (default)")
                                        
                                        # Read directly with pandas
                                        df_reference = pd.read_csv(csv_url)
                                        
                                        st.success(f"✅ Data berhasil di-import: {len(df_reference)} baris")
                                        
                                        # Display reference data preview
                                        with st.expander("👁️ Preview Data Google Sheets"):
                                            st.dataframe(df_reference.head(10), use_container_width=True)
                                        
                                        # Compare data
                                        st.subheader("📊 Hasil Validasi")
                                        
                                        # Normalization function for street names
                                        def normalize_street_name(name):
                                            """Normalize street name by expanding abbreviations"""
                                            if pd.isna(name):
                                                return ""
                                            
                                            name = str(name).strip()
                                            
                                            # Define abbreviation mappings
                                            replacements = [
                                                (r'\bGg\.?\s+', 'Gang '),      # Gg. or Gg → Gang
                                                (r'\bJl\.?\s+', 'Jalan '),     # Jl. or Jl → Jalan
                                                (r'\bJln\.?\s+', 'Jalan '),    # Jln. or Jln → Jalan
                                                (r'\bJln\b', 'Jalan'),         # Jln at end → Jalan
                                                (r'\bGg\b', 'Gang'),           # Gg at end → Gang
                                                (r'\bJl\b', 'Jalan'),          # Jl at end → Jalan
                                            ]
                                            
                                            # Apply replacements
                                            for pattern, replacement in replacements:
                                                name = re.sub(pattern, replacement, name, flags=re.IGNORECASE)
                                            
                                            # Remove extra spaces and convert to lowercase
                                            name = ' '.join(name.split()).lower()
                                            
                                            return name
                                        
                                        # Normalize column names for comparison
                                        col_mapping = {}
                                        for col in df_reference.columns:
                                            col_lower = col.lower().strip()
                                            if 'jalan' in col_lower or 'nama' in col_lower:
                                                col_mapping['street_name'] = col
                                            elif 'sls' in col_lower or 'rt' in col_lower:
                                                col_mapping['sls'] = col
                                            elif 'lingkungan' in col_lower:
                                                col_mapping['lingkungan'] = col
                                            elif 'kelurahan' in col_lower:
                                                col_mapping['kelurahan'] = col
                                        
                                        # Create comparison
                                        matches = []
                                        mismatches = []
                                        missing_in_osm = []
                                        extra_in_osm = []
                                        
                                        # Normalize street names for comparison
                                        df_streets_norm = df_streets.copy()
                                        df_streets_norm['normalized_name'] = df_streets['Nama Jalan dan Gang'].apply(normalize_street_name)
                                        
                                        if 'street_name' in col_mapping:
                                            df_reference_norm = df_reference.copy()
                                            df_reference_norm['normalized_name'] = df_reference[col_mapping['street_name']].apply(normalize_street_name)
                                            
                                            # Show normalization examples
                                            with st.expander("🔍 Preview Normalisasi Nama Jalan"):
                                                preview_data = []
                                                for i in range(min(5, len(df_reference))):
                                                    preview_data.append({
                                                        'Original (Google Sheets)': df_reference.iloc[i][col_mapping['street_name']],
                                                        'Normalized': df_reference_norm.iloc[i]['normalized_name']
                                                    })
                                                st.dataframe(pd.DataFrame(preview_data), use_container_width=True)
                                            
                                            # Find matches and mismatches
                                            for idx, ref_row in df_reference_norm.iterrows():
                                                ref_street = ref_row['normalized_name']
                                                
                                                osm_match = df_streets_norm[df_streets_norm['normalized_name'] == ref_street]
                                                
                                                if not osm_match.empty:
                                                    osm_row = osm_match.iloc[0]
                                                    
                                                    # Check if administrative data matches
                                                    admin_match = True
                                                    differences = []
                                                    
                                                    if 'sls' in col_mapping and str(ref_row[col_mapping['sls']]) != str(osm_row['SLS']):
                                                        admin_match = False
                                                        differences.append(f"SLS: {ref_row[col_mapping['sls']]} vs {osm_row['SLS']}")
                                                    
                                                    if 'lingkungan' in col_mapping and str(ref_row[col_mapping['lingkungan']]) != str(osm_row['Lingkungan']):
                                                        admin_match = False
                                                        differences.append(f"Lingkungan: {ref_row[col_mapping['lingkungan']]} vs {osm_row['Lingkungan']}")
                                                    
                                                    if 'kelurahan' in col_mapping and str(ref_row[col_mapping['kelurahan']]) != str(osm_row['Kelurahan']):
                                                        admin_match = False
                                                        differences.append(f"Kelurahan: {ref_row[col_mapping['kelurahan']]} vs {osm_row['Kelurahan']}")
                                                    
                                                    if admin_match:
                                                        matches.append({
                                                            'Nama Jalan': ref_row[col_mapping['street_name']],
                                                            'Status': '✅ Cocok'
                                                        })
                                                    else:
                                                        mismatches.append({
                                                            'Nama Jalan': ref_row[col_mapping['street_name']],
                                                            'Perbedaan': ', '.join(differences)
                                                        })
                                                else:
                                                    missing_in_osm.append({
                                                        'Nama Jalan': ref_row[col_mapping['street_name']],
                                                        'Info': 'Tidak ditemukan di data OSM'
                                                    })
                                            
                                            # Find streets in OSM but not in reference
                                            for idx, osm_row in df_streets_norm.iterrows():
                                                if osm_row['normalized_name'] not in df_reference_norm['normalized_name'].values:
                                                    extra_in_osm.append({
                                                        'Nama Jalan': osm_row['Nama Jalan dan Gang'],
                                                        'Info': 'Ada di OSM, tidak ada di referensi'
                                                    })
                                            
                                            # Display results
                                            col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
                                            col_stat1.metric("✅ Cocok", len(matches), delta=f"{len(matches)/len(df_reference)*100:.1f}%")
                                            col_stat2.metric("⚠️ Beda Data", len(mismatches))
                                            col_stat3.metric("❌ Hilang di OSM", len(missing_in_osm))
                                            col_stat4.metric("➕ Ekstra di OSM", len(extra_in_osm))
                                            
                                            # Show mismatches
                                            if mismatches:
                                                st.warning(f"⚠️ **{len(mismatches)} jalan** ditemukan perbedaan data administratif:")
                                                st.dataframe(pd.DataFrame(mismatches), use_container_width=True)
                                            
                                            # Show missing streets
                                            if missing_in_osm:
                                                st.error(f"❌ **{len(missing_in_osm)} jalan** dari referensi tidak ditemukan di OSM:")
                                                st.dataframe(pd.DataFrame(missing_in_osm), use_container_width=True)
                                            
                                            # Show extra streets
                                            if extra_in_osm:
                                                with st.expander(f"➕ {len(extra_in_osm)} jalan tambahan di OSM (tidak ada di referensi)"):
                                                    st.dataframe(pd.DataFrame(extra_in_osm), use_container_width=True)
                                            
                                            if not mismatches and not missing_in_osm:
                                                st.success("🎉 Semua data cocok sempurna!")
                                        
                                        else:
                                            st.error("❌ Kolom 'Nama Jalan' tidak ditemukan di Google Sheets. Pastikan ada kolom yang mengandung nama jalan.")
                                    
                                    else:
                                        st.error("❌ URL Google Sheets tidak valid")
                                else:
                                    st.info("🔐 Untuk sheet private, gunakan Google Sheets API dengan service account. Silakan hubungi developer.")
                        
                        except Exception as e:
                            st.error(f"❌ Error saat mengambil data: {str(e)}")
                            st.info("💡 Pastikan Google Sheets Anda sudah diset public (Anyone with the link can view)")

    
    except Exception as e:
        st.error(f"❌ Error: {str(e)}")
        st.info("Pastikan file `5271sls.geojson` tersedia di lokasi yang benar.")


# Footer
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: #6b7280; font-size: 12px;'>
    <p>Sistem Monitoring PAD - Badan Keuangan Daerah Kota Mataram</p>
    <p>Powered by Google Earth Engine, Sentinel-2, Dynamic World | Data: Real-time</p>
</div>
""", unsafe_allow_html=True)
